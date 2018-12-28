import pymysql.cursors
from openpyxl import Workbook
from openpyxl.styles import Alignment, Fill
import re
import datetime
import sys

paramCheck = True
if len(sys.argv) == 1:
    month = datetime.datetime.now().month
    year = datetime.datetime.now().year
elif len(sys.argv) == 2:
    try:
        month = int(sys.argv[1])
    except ValueError:
        print('Month incorrect')
        paramCheck = False
    else:
        if month < 0 or month > 12:
            print('Month incorrect')
            paramCheck = False
        else:
            year = datetime.datetime.now().year
else:
    try:
        month = int(sys.argv[1])
        year = int(sys.argv[2])
    except ValueError:
        print('Month or Year incorrect')
        paramCheck = False
    else:
        if month < 0 or month > 12 or year < 0 or year > 2100:
            print('Month or Year incorrect')
            paramCheck = False

if paramCheck:
    if month in [1, 3, 5, 7, 8, 10, 12]:
        timeReportStart = '{}-{}-1'.format(year, month)
        timeReportEnd = '{}-{}-31 23:59:59'.format(year, month)
    elif month == 2:
        timeReportStart = '{}-2-1'.format(year)
        if (year%4) == 0:
            timeReportEnd = '{}-2-29 23:59:59'.format(year)
        else:
            timeReportEnd = '{}-2-28 23:59:59'.format(year)
    else:
        timeReportStart = '{}-{}-1'.format(year, month)
        timeReportEnd = '{}-{}-30 23:59:59'.format(year, month)
    # Connect to the database
    connection = pymysql.connect(host='192.168.100.39',
                                 user='thaonv',
                                 password='meditech2017',
                                 db='nagios',
                                 charset='utf8mb4',
                                 cursorclass=pymysql.cursors.DictCursor)

    workbook = Workbook()
    excelFilename = 'Bao_cao_chi_tieu_KPI_server_{}_{}.xlsx'.format(month, year)
    sheet1 = workbook.active
    sheet1.title = "active"
    centerAlignment = Alignment(horizontal="center", vertical="center")
    # greenFill = Fill("solid", fgColor="7CFC00")
    # orangeFill = Fill("solid", fgColor="FFA500")
    sheet1RowCursor = 3

    sheet1.merge_cells(start_row=1, end_row=2, start_column=1, end_column=1)
    cell = sheet1.cell(1, 1, 'IP address')
    cell.alignment = centerAlignment

    sheet1.merge_cells(start_row=1, end_row=1, start_column=2, end_column=4)
    cell = sheet1.cell(1, 2, 'Avaibility')
    cell.alignment = centerAlignment
    sheet1.cell(2, 2, 'min')
    sheet1.cell(2, 3, 'max')
    sheet1.cell(2, 4, 'average')

    sheet1.merge_cells(start_row=1, end_row=1, start_column=5, end_column=7)
    cell = sheet1.cell(1, 5, 'CPU Utilization')
    cell.alignment = centerAlignment
    sheet1.cell(2, 5, 'min')
    sheet1.cell(2, 6, 'max')
    sheet1.cell(2, 7, 'average')

    sheet1.merge_cells(start_row=1, end_row=1, start_column=8, end_column=10)
    cell = sheet1.cell(1, 8, 'Memory Used')
    cell.alignment = centerAlignment
    sheet1.cell(2, 8, 'min')
    sheet1.cell(2, 9, 'max')
    sheet1.cell(2, 10, 'average')

    sheet1.merge_cells(start_row=1, end_row=2, start_column=11, end_column=11)
    cell = sheet1.cell(1, 11, 'Giảm trừ hợp lệ')
    cell.alignment = centerAlignment

    sheet1.merge_cells(start_row=1, end_row=2, start_column=12, end_column=12)
    cell = sheet1.cell(1, 12, 'Thời gian downtime')
    cell.alignment = centerAlignment



    try:
        # get host group
        with connection.cursor() as cursor:
            sql = "SELECT host_object_id, display_name, address FROM nagios_hosts"
            cursor.execute(sql)
            hostGroup = cursor.fetchall()

        # get CPU group
        with connection.cursor() as cursor:
            sql = "SELECT host_object_id, service_object_id FROM nagios_services WHERE display_name='CPU utilization'"
            cursor.execute(sql)
            cpuGroup = cursor.fetchall()

        # get memory group
        with connection.cursor() as cursor:
            sql = "SELECT host_object_id, service_object_id FROM nagios_services WHERE display_name IN ('Memory', 'Memory used')"
            cursor.execute(sql)
            memoryGroup = cursor.fetchall()

        for hostIndex, hostItem in enumerate(hostGroup):
            sheet1.cell(hostIndex+sheet1RowCursor, 1, hostItem['address'])
            with connection.cursor() as cursor:
                sql = "SELECT output, start_time FROM nagios_hostchecks WHERE host_object_id={} and start_time between '{}' and '{}'".format(hostItem['host_object_id'], timeReportStart, timeReportEnd)
                cursor.execute(sql)
                downtime = datetime.timedelta(seconds=0)
                availabilityDetail = {}
                availabilitySum = 0
                availabilityItem = cursor.rowcount
                previosStatus = True
                previosTime = None
                if availabilityItem > 0:
                    for cursorIndex, cursorItem in enumerate(cursor.fetchall()):
                        if ('lost 100%' or 'Host Check Timed Out') in cursorItem['output']:
                            if not previosStatus:
                                downtime += cursorItem['start_time'] - previosTime
                            previosStatus = False
                            previosTime = cursorItem['start_time']
                        else:
                            previosStatus = True
                            availabilitySum += 100
                    downtime = round(downtime.total_seconds()/60)
                    if downtime == 0:
                        availabilityDetail['min'] = 100
                        availabilityDetail['average'] = 100
                    else:
                        availabilityDetail['min'] = 0
                        availabilityDetail['average'] = round((availabilitySum / availabilityItem), 2)
                        sheet1.cell(hostIndex + sheet1RowCursor, 12, '{} phút'.format(downtime))
                    availabilityDetail['max'] = 100

                    sheet1.cell(hostIndex + sheet1RowCursor, 2, availabilityDetail['min'])
                    sheet1.cell(hostIndex + sheet1RowCursor, 3, availabilityDetail['max'])
                    cell = sheet1.cell(hostIndex + sheet1RowCursor, 4, availabilityDetail['average'])

            for cpuItem in cpuGroup:
                cpuObjectId = None
                if cpuItem['host_object_id'] == hostItem['host_object_id']:
                    cpuObjectId = cpuItem['service_object_id']
                if cpuObjectId is not None:
                    with connection.cursor() as cursor:
                        sql = "SELECT output FROM nagios_servicechecks WHERE service_object_id={} and start_time between '{}' and '{}'".format(cpuObjectId, timeReportStart, timeReportEnd)
                        cursor.execute(sql)
                        cpuDetail = {}
                        cpuSum = float(0)
                        cpuNumberItem = cursor.rowcount
                        if cpuNumberItem > 0:
                            for cursorIndex, cursorItem in enumerate(cursor.fetchall()):
                                if 'total' in cursorItem['output']:
                                    cpuUsed = float(re.findall(r"total: ([\d.]*)%", cursorItem['output'])[0])
                                else:
                                    cpuUsed = float(re.findall(r"OK - ([\d.]*)%", cursorItem['output'])[0])
                                if cursorIndex == 0:
                                    cpuDetail['min'] = cpuUsed
                                    cpuDetail['max'] = cpuUsed
                                else:
                                    cpuDetail['min'] = min(cpuUsed, cpuDetail['min'])
                                    cpuDetail['max'] = max(cpuUsed, cpuDetail['max'])
                                    cpuSum += cpuUsed
                            cpuDetail['min'] = round(cpuDetail['min'], 2)
                            cpuDetail['max'] = round(cpuDetail['max'], 2)
                            cpuDetail['average'] = round((cpuSum / cpuNumberItem), 2)

                            sheet1.cell(hostIndex + sheet1RowCursor, 5, cpuDetail['min'])
                            sheet1.cell(hostIndex + sheet1RowCursor, 6, cpuDetail['max'])
                            sheet1.cell(hostIndex + sheet1RowCursor, 7, cpuDetail['average'])

            for memoryItem in memoryGroup:
                memoryObjectId = None
                if memoryItem['host_object_id'] == hostItem['host_object_id']:
                    memoryObjectId = memoryItem['service_object_id']
                if memoryObjectId is not None:
                    with connection.cursor() as cursor:
                        sql = "SELECT output FROM nagios_servicechecks WHERE service_object_id={} and start_time between '{}' and '{}'".format(memoryObjectId, timeReportStart, timeReportEnd)
                        cursor.execute(sql)
                        memoryDetail = {}
                        memorySum = float(0)
                        memoryNumberItem = cursor.rowcount
                        if memoryNumberItem > 0:
                            for cursorIndex, cursorItem in enumerate(cursor.fetchall()):
                                memoryUsed = re.findall(r"OK - ([\d.]*)%", cursorItem['output'])
                                if len(memoryUsed) == 0:
                                    if 'RAM used' in cursorItem['output']:
                                        memoryUsed = float(re.findall(r"GB \(([\d.]*)%\)", cursorItem['output'])[0])
                                    else:
                                        break
                                else:
                                    memoryUsed = float(memoryUsed[0])
                                if cursorIndex == 0:
                                    memoryDetail['min'] = memoryUsed
                                    memoryDetail['max'] = memoryUsed
                                else:
                                    memoryDetail['min'] = min(memoryUsed, memoryDetail['min'])
                                    memoryDetail['max'] = max(memoryUsed, memoryDetail['max'])
                                    memorySum += memoryUsed
                            memoryDetail['min'] = round(memoryDetail['min'], 2)
                            memoryDetail['max'] = round(memoryDetail['max'], 2)
                            memoryDetail['average'] = round((memorySum / memoryNumberItem), 2)
                            sheet1.cell(hostIndex + sheet1RowCursor, 8, memoryDetail['min'])
                            sheet1.cell(hostIndex + sheet1RowCursor, 9, memoryDetail['max'])
                            sheet1.cell(hostIndex + sheet1RowCursor, 10, memoryDetail['average'])
        workbook.save(filename=excelFilename)
    finally:
        connection.close()
